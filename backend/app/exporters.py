"""把 reader 已经聚合出的数据转换成 CSV/XLSX 字节流。

设计要点：
- 三种导出场景共享一组小帮助函数（_rows_to_csv / _rows_to_xlsx）。
- 多表导出时 CSV 用「文件名 -> bytes」字典 + 一个 zip 打包；XLSX 用多 sheet。
- 调用方只需选 format 并提供数据；本模块不直接访问 SQLite，便于单测。
"""
from __future__ import annotations

import csv
import io
import zipfile
from collections.abc import Iterable, Sequence
from typing import Any

from openpyxl import Workbook


def _stringify(value: Any) -> Any:
    """让 csv/xlsx 能直接写入：dict/list 转 JSON 字符串，None 转空。"""
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        import json

        return json.dumps(value, ensure_ascii=False)
    return value


def rows_to_csv_bytes(
    rows: Sequence[dict[str, Any]],
    fieldnames: Sequence[str] | None = None,
) -> bytes:
    if not rows:
        return ("" if fieldnames is None else ",".join(fieldnames) + "\r\n").encode(
            "utf-8-sig"
        )
    if fieldnames is None:
        fieldnames = list(rows[0].keys())
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(fieldnames), extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({k: _stringify(row.get(k)) for k in fieldnames})
    # 加 BOM 让 Excel 直接识别 UTF-8
    return buf.getvalue().encode("utf-8-sig")


def sheets_to_xlsx_bytes(
    sheets: Iterable[tuple[str, Sequence[dict[str, Any]], Sequence[str] | None]],
) -> bytes:
    """sheets: iterable of (sheet_name, rows, fieldnames or None)."""
    wb = Workbook()
    # 默认有个空 sheet，先 remove
    default = wb.active
    wb.remove(default)
    any_sheet = False
    for sheet_name, rows, fieldnames in sheets:
        # Excel sheet 名长度上限 31 + 不允许某些字符
        safe_name = (sheet_name or "sheet")[:31]
        for ch in "[]:*?/\\":
            safe_name = safe_name.replace(ch, "_")
        ws = wb.create_sheet(safe_name)
        any_sheet = True
        if rows:
            fields = list(fieldnames) if fieldnames else list(rows[0].keys())
        else:
            fields = list(fieldnames) if fieldnames else []
        if fields:
            ws.append(fields)
            for row in rows:
                ws.append([_stringify(row.get(k)) for k in fields])
    if not any_sheet:
        wb.create_sheet("empty")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def csv_files_to_zip_bytes(files: dict[str, bytes]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, content in files.items():
            zf.writestr(name, content)
    return buf.getvalue()


# ---------- 业务层组装 ----------


def export_run_results(run_payload: dict[str, Any], fmt: str) -> tuple[bytes, str, str]:
    """整批结果导出：labels / evidence / coverage / failures / features 几张表。

    run_payload 期望来自 LabelRunReader.get_run_export(run_id) ——
    各张表 list[dict] 集中在一处。
    """
    fmt = fmt.lower()
    sheets_def: list[tuple[str, list[dict[str, Any]], list[str] | None]] = [
        (
            "labels",
            run_payload["labels"],
            [
                "fund_code",
                "label_code",
                "label_name",
                "category",
                "status",
                "confidence",
            ],
        ),
        (
            "evidence",
            run_payload["evidence"],
            ["fund_code", "label_code", "metric", "value", "threshold", "source", "message"],
        ),
        (
            "coverage",
            run_payload["coverage"],
            ["fund_code", "missing_fields", "review_action"],
        ),
        (
            "failures",
            run_payload["failures"],
            ["fund_code", "stage", "error_type", "message", "recorded_at"],
        ),
        (
            "features",
            run_payload["features"],
            ["fund_code", "feature_code", "value", "source"],
        ),
        (
            "portfolio_matrix",
            run_payload.get("portfolio_matrix", []),
            [
                "fund_code",
                "allocation_status",
                "portfolio_roles",
                "style_tags",
                "return_tags",
                "risk_tags",
                "data_tags",
                "blocking_reasons",
                "watch_reasons",
                "alpha_1y",
                "information_ratio_1y",
                "annualized_excess_return_1y",
                "max_drawdown_1y",
                "annualized_volatility_1y",
                "sharpe_ratio_1y",
                "beta_1y",
                "quality_growth_weight",
                "deep_value_weight",
                "dividend_steady_weight",
                "label_codes",
                "group_codes",
                "classifications",
            ],
        ),
        (
            "portfolio_draft",
            run_payload.get("portfolio_draft", []),
            [
                "fund_code",
                "bucket",
                "draft_weight_pct",
                "max_weight_pct",
                "score",
                "portfolio_roles",
                "risk_tags",
            ],
        ),
        (
            "factor_exposures",
            run_payload.get("factor_exposures", []),
            [
                "fund_code",
                "report_date",
                "factor_code",
                "exposure_value",
                "coverage_weight",
                "holding_total_weight",
                "stock_count",
                "covered_stock_count",
                "source",
                "as_of_date",
                "computed_at",
            ],
        ),
        (
            "equity_style_contributions",
            run_payload.get("equity_style_contributions", []),
            [
                "fund_code",
                "report_date",
                "stock_code",
                "stock_name",
                "weight",
                "style_code",
                "style_name",
                "matched",
                "contribution_weight",
                "factor_values_json",
                "rule_snapshot_json",
                "factor_as_of_date",
                "source",
                "computed_at",
            ],
        ),
        (
            "calculations",
            run_payload["calculations"],
            [
                "fund_code",
                "label_code",
                "label_name",
                "category",
                "state",
                "reason_code",
                "observed",
                "threshold",
                "source",
                "message",
            ],
        ),
        (
            "classifications",
            run_payload["classifications"],
            [
                "fund_code",
                "dimension",
                "classification_code",
                "classification_name",
                "confidence",
                "reason_code",
                "evidence",
                "source",
            ],
        ),
        (
            "groups",
            run_payload["groups"],
            [
                "fund_code",
                "group_code",
                "group_name",
                "group_type",
                "reason_code",
                "evidence",
                "source",
            ],
        ),
    ]
    if fmt == "xlsx":
        data = sheets_to_xlsx_bytes(sheets_def)
        return (
            data,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"run_{run_payload['run_id']}.xlsx",
        )
    if fmt == "csv":
        files = {
            f"{name}.csv": rows_to_csv_bytes(rows, fields)
            for name, rows, fields in sheets_def
        }
        return (
            csv_files_to_zip_bytes(files),
            "application/zip",
            f"run_{run_payload['run_id']}.zip",
        )
    raise ValueError(f"unsupported format: {fmt}")


def export_cognition_result(
    result: dict[str, Any], fmt: str
) -> tuple[bytes, str, str]:
    """导出认知引擎结果为 CSV/XLSX。

    包含 4 个 sheet：认知摘要、匹配基金、组合方案、组合风险指标。
    """
    direction = result.get("direction", result.get("stock_name", "cognition"))
    conviction = result.get("conviction", "medium")
    base = f"cognition_{direction}"

    # Sheet 1: 认知摘要
    summary_rows = [{
        "direction": direction,
        "conviction": conviction,
        "belief": (result.get("step1_judgment") or {}).get("belief", ""),
        "available_links": "、".join(result.get("available_links", [])),
        "fund_count": len(result.get("step4_fund_matches", [])),
    }]

    # Sheet 2: 匹配基金
    fund_rows = []
    for f in result.get("step4_fund_matches", []):
        val = f.get("valuation", {})
        gate = f.get("gate", {})
        fund_rows.append({
            "fund_code": f.get("fund_code", ""),
            "fund_name": f.get("fund_name", ""),
            "match_pct": f.get("match_pct", ""),
            "weighted_pe": val.get("weighted_pe", ""),
            "weighted_val_pct": val.get("weighted_val_pct", ""),
            "peg": val.get("peg", ""),
            "price_in_years": val.get("price_in_years", ""),
            "val_judge": val.get("val_judge", ""),
            "gate_passed": gate.get("passed", ""),
            "gate_violations": "; ".join(gate.get("violations", [])),
            "trend": (f.get("trend") or {}).get("trend", ""),
        })

    # Sheet 3: 组合方案
    portfolio = result.get("step5_portfolio", {})
    pf_rows = []
    for f in portfolio.get("selected_funds", []):
        val = f.get("valuation", {})
        pf_rows.append({
            "fund_code": f.get("fund_code", ""),
            "fund_name": f.get("fund_name", ""),
            "weight": f.get("weight", ""),
            "match_pct": f.get("match_pct", ""),
            "max_weight": f.get("max_weight", ""),
            "weighted_pe": val.get("weighted_pe", ""),
            "val_judge": val.get("val_judge", ""),
        })
    if portfolio.get("defense_position"):
        df = portfolio["defense_position"]
        val = df.get("valuation", {})
        pf_rows.append({
            "fund_code": df.get("fund_code", ""),
            "fund_name": df.get("fund_name", ""),
            "weight": df.get("weight", ""),
            "match_pct": df.get("match_pct", ""),
            "max_weight": "",
            "weighted_pe": val.get("weighted_pe", ""),
            "val_judge": val.get("val_judge", ""),
        })

    # Sheet 4: 组合风险指标
    metrics = portfolio.get("metrics", {})
    metrics_rows = [
        {"指标": "组合加权PE", "值": metrics.get("portfolio_pe", "-")},
        {"指标": "年化波动率(%)", "值": metrics.get("portfolio_volatility", "-")},
        {"指标": "最大回撤(%)", "值": metrics.get("portfolio_max_drawdown", "-")},
        {"指标": "现金比例(%)", "值": portfolio.get("cash_pct", "-")},
        {"指标": "总投资比例(%)", "值": portfolio.get("total_invested", "-")},
    ]
    for h in metrics.get("holdings_penetration", []):
        metrics_rows.append({
            "指标": f"持仓-{h.get('stock_name', h.get('stock_code', ''))}",
            "值": h.get("weight", "-"),
        })
    for ind in metrics.get("industry_exposure", []):
        metrics_rows.append({
            "指标": f"行业-{ind.get('name', '')}",
            "值": ind.get("weight", "-"),
        })

    summary_fields = ["direction", "conviction", "belief", "available_links", "fund_count"]
    fund_fields = [
        "fund_code", "fund_name", "match_pct", "weighted_pe",
        "weighted_val_pct", "peg", "price_in_years", "val_judge",
        "gate_passed", "gate_violations", "trend",
    ]
    pf_fields = ["fund_code", "fund_name", "weight", "match_pct", "max_weight", "weighted_pe", "val_judge"]
    metrics_fields = ["指标", "值"]

    sheets = [
        ("认知摘要", summary_rows, summary_fields),
        ("匹配基金", fund_rows, fund_fields),
        ("组合方案", pf_rows, pf_fields),
        ("组合风险指标", metrics_rows, metrics_fields),
    ]

    if fmt == "csv":
        files: dict[str, bytes] = {}
        for name, rows, fields in sheets:
            files[f"{name}.csv"] = rows_to_csv_bytes(rows, fields)
        return (
            csv_files_to_zip_bytes(files),
            "application/zip",
            f"{base}.zip",
        )

    return (
        sheets_to_xlsx_bytes(sheets),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"{base}.xlsx",
    )


def export_fund_report(report: dict[str, Any], fmt: str) -> tuple[bytes, str, str]:
    """单基金 report 导出（labels/evidence/features/coverage/reviews）。"""
    fmt = fmt.lower()
    sheets_def: list[tuple[str, list[dict[str, Any]], list[str] | None]] = [
        (
            "summary",
            [report.get("summary", {})],
            None,
        ),
        ("labels", report.get("labels", []), None),
        ("evidence", report.get("evidence", []), None),
        ("features", report.get("features", []), None),
        ("factor_exposures", report.get("factor_exposures", []), None),
        (
            "equity_style_contributions",
            report.get("equity_style_contributions", []),
            None,
        ),
        ("calculations", report.get("calculations", []), None),
        ("classifications", report.get("classifications", []), None),
        ("groups", report.get("groups", []), None),
        ("coverage", [report.get("coverage", {})] if report.get("coverage") else [], None),
        ("reviews", report.get("reviews", []), None),
    ]
    base = f"fund_{report.get('fund_code', 'unknown')}_run_{report.get('run_id', 'unknown')}"
    if fmt == "xlsx":
        return (
            sheets_to_xlsx_bytes(sheets_def),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"{base}.xlsx",
        )
    if fmt == "csv":
        files = {
            f"{name}.csv": rows_to_csv_bytes(rows, fields)
            for name, rows, fields in sheets_def
        }
        return (
            csv_files_to_zip_bytes(files),
            "application/zip",
            f"{base}.zip",
        )
    raise ValueError(f"unsupported format: {fmt}")


def export_review_queue(
    rows: list[dict[str, Any]], run_id: str, fmt: str
) -> tuple[bytes, str, str]:
    fmt = fmt.lower()
    fields = ["fund_code", "label_count", "review_action", "missing_field_count"]
    base = f"review_queue_{run_id}"
    if fmt == "csv":
        return (
            rows_to_csv_bytes(rows, fields),
            "text/csv; charset=utf-8",
            f"{base}.csv",
        )
    if fmt == "xlsx":
        return (
            sheets_to_xlsx_bytes([("review_queue", rows, fields)]),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"{base}.xlsx",
        )
    raise ValueError(f"unsupported format: {fmt}")
