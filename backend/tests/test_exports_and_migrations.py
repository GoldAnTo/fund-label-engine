"""导出 / 迁移 / Phase 5 schema 的回归测试。"""
from __future__ import annotations

import csv
import io
import sqlite3
import zipfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.batch import run_batch
from app.main import create_app
from app.persistence import LabelRunReader
from app.persistence.migrations_runner import run_migrations


@pytest.fixture()
def seeded_db(tmp_path: Path) -> Path:
    from scripts.seed_sample_db import seed

    db = tmp_path / "fund.sqlite"
    seed(db)
    return db


@pytest.fixture()
def seeded_run(seeded_db: Path) -> tuple[Path, str]:
    run_id, _ = run_batch(seeded_db)
    return seeded_db, run_id


def test_run_export_csv_returns_zip_with_expected_files(seeded_run) -> None:
    db, run_id = seeded_run
    client = TestClient(create_app(db_path=db))

    resp = client.get(f"/v1/runs/{run_id}/export", params={"format": "csv"})

    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/zip"
    zf = zipfile.ZipFile(io.BytesIO(resp.content))
    names = set(zf.namelist())
    assert {
        "labels.csv",
        "evidence.csv",
        "coverage.csv",
        "failures.csv",
        "features.csv",
        "portfolio_matrix.csv",
        "calculations.csv",
        "classifications.csv",
        "groups.csv",
    } <= names
    # labels.csv 应该至少有 fund_code + label_code 两列
    with zf.open("labels.csv") as f:
        text = f.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    assert "fund_code" in reader.fieldnames
    assert "label_code" in reader.fieldnames
    rows = list(reader)
    assert len(rows) > 0


def test_run_export_xlsx_has_all_sheets(seeded_run) -> None:
    db, run_id = seeded_run
    client = TestClient(create_app(db_path=db))

    resp = client.get(f"/v1/runs/{run_id}/export", params={"format": "xlsx"})

    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    wb = load_workbook(io.BytesIO(resp.content))
    assert {
        "labels",
        "evidence",
        "coverage",
        "failures",
        "features",
        "portfolio_matrix",
        "calculations",
        "classifications",
        "groups",
    } <= set(wb.sheetnames)
    labels_sheet = wb["labels"]
    header = [c.value for c in labels_sheet[1]]
    assert "fund_code" in header and "label_code" in header


def test_fund_report_export_csv_contains_summary_and_labels(seeded_run) -> None:
    db, run_id = seeded_run
    client = TestClient(create_app(db_path=db))

    resp = client.get(
        f"/v1/runs/{run_id}/funds/000001/export", params={"format": "csv"}
    )

    assert resp.status_code == 200
    zf = zipfile.ZipFile(io.BytesIO(resp.content))
    assert "summary.csv" in zf.namelist()
    assert "labels.csv" in zf.namelist()
    assert "evidence.csv" in zf.namelist()


def test_review_queue_export_csv_returns_plain_csv(seeded_run) -> None:
    db, run_id = seeded_run
    client = TestClient(create_app(db_path=db))

    resp = client.get(
        f"/v1/runs/{run_id}/review-queue/export", params={"format": "csv"}
    )

    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    text = resp.content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    assert "fund_code" in reader.fieldnames
    assert "review_action" in reader.fieldnames


def test_export_unknown_run_returns_404(seeded_db: Path) -> None:
    client = TestClient(create_app(db_path=seeded_db))
    resp = client.get("/v1/runs/does-not-exist/export")
    assert resp.status_code == 404


def test_migrations_run_idempotently_on_fresh_db(tmp_path: Path) -> None:
    db = tmp_path / "fresh.sqlite"
    # 给 migration 一个目标：先建好 label_runs / label_definitions 基线表
    sqlite3.connect(db).executescript(
        """
        CREATE TABLE label_runs (run_id TEXT PRIMARY KEY);
        CREATE TABLE label_definitions (
            label_code TEXT, rule_version TEXT,
            PRIMARY KEY (label_code, rule_version)
        );
        """
    )

    executed = run_migrations(str(db))
    assert "0001_baseline" in executed
    assert "0003_phase5_stock_factors_and_labels" in executed
    # 第二次运行返回空（幂等）
    again = run_migrations(str(db))
    assert again == []


def test_migrations_create_phase5_tables(seeded_run) -> None:
    db, _ = seeded_run
    with sqlite3.connect(db) as conn:
        tables = {
            row[0]
            for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            )
        }
    assert "stock_factor_values" in tables
    assert "stock_labels" in tables
    assert "label_calculation_states" in tables
    assert "fund_classification_results" in tables
    assert "fund_group_results" in tables
    assert "schema_migrations" in tables


def test_migrations_create_stock_industry_map_table(tmp_path: Path) -> None:
    db = tmp_path / "industry-map.sqlite"
    # migration 0002/0004 会 ALTER 基线表，需先建 label_runs / label_definitions
    sqlite3.connect(db).executescript(
        """
        CREATE TABLE label_runs (run_id TEXT PRIMARY KEY);
        CREATE TABLE label_definitions (
            label_code TEXT, rule_version TEXT,
            PRIMARY KEY (label_code, rule_version)
        );
        """
    )
    run_migrations(str(db))

    with sqlite3.connect(db) as conn:
        tables = {
            row[0]
            for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        }
        cols = {
            row[1]
            for row in conn.execute("PRAGMA table_info(stock_industry_map)").fetchall()
        }

    assert "stock_industry_map" in tables
    assert {
        "stock_code",
        "industry_code",
        "industry_name",
        "sector_group",
        "source",
        "as_of_date",
    }.issubset(cols)


def test_migrations_create_equity_style_contributions_table(seeded_run) -> None:
    db, _ = seeded_run
    with sqlite3.connect(db) as conn:
        tables = {
            row[0]
            for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            )
        }
    assert "fund_equity_style_contributions" in tables


def test_reader_fund_report_includes_label_calculation_states(seeded_run) -> None:
    db, run_id = seeded_run

    report = LabelRunReader(db).get_fund_report(run_id, "000001")

    assert report is not None
    calculations = {item["label_code"]: item for item in report["calculations"]}
    assert calculations["holding_concentration_high"]["state"] == "triggered"
    assert calculations["fund_size_small"]["state"] == "not_triggered"
    assert calculations["long_term_return_strong"]["state"] == "not_computed"
    assert calculations["long_term_return_strong"]["reason_code"] == "return_window_insufficient"
    classifications = {
        item["dimension"]: item for item in report["classifications"]
    }
    assert classifications["asset_class"]["classification_code"] == "equity_related"
    assert report["groups"]
    assert report["summary"]["classification_count"] == len(report["classifications"])
    assert report["summary"]["group_count"] == len(report["groups"])


def test_reader_stock_factor_methods_return_empty_when_no_data(seeded_run) -> None:
    db, _ = seeded_run
    reader = LabelRunReader(db)
    assert reader.list_stock_factors() == []
    assert reader.list_stock_labels(stock_code="600519") == []


def test_reader_stock_factor_returns_rows_after_insert(seeded_run) -> None:
    db, _ = seeded_run
    with sqlite3.connect(db) as conn:
        conn.execute(
            "INSERT INTO stock_factor_values VALUES (?,?,?,?,?)",
            ("600519", "pb", 8.5, "2025-06-30", "fundamentals"),
        )
        conn.commit()
    rows = LabelRunReader(db).list_stock_factors(stock_code="600519")
    assert len(rows) == 1
    assert rows[0]["factor_code"] == "pb"
    assert rows[0]["factor_value"] == 8.5
